"""Test per GET /api/v1/raw/excel/template.

Copertura:
1. Editor scarica → 200 + content-type xlsx
2. Viewer → 403
3. Workbook ha esattamente 3 fogli scopeX + 1 _README
4. Ogni foglio scope ha gli header attesi (riga 1 match alle colonne canoniche)
5. Header è in bold (verifica via openpyxl)
"""

from __future__ import annotations

import io
import os
import uuid
from typing import Any
from unittest.mock import AsyncMock, MagicMock

# Configura variabili d'ambiente prima degli import del progetto
os.environ.setdefault("GHG_JWT_ALGORITHM", "HS256")
os.environ.setdefault("GHG_JWT_SECRET", "test-secret-key-for-unit-tests-only")
os.environ.setdefault("GHG_ENVIRONMENT", "development")

import openpyxl
from fastapi.testclient import TestClient

from ghg_tool.api.dependencies.auth import CurrentUser, get_current_user
from ghg_tool.api.dependencies.db import get_db
from ghg_tool.api.main import app
from ghg_tool.etl.builders.excel_template import (
    _SCOPE1_HEADERS,
    _SCOPE2_HEADERS,
    _SCOPE3_HEADERS,
    build_excel_template,
)

# ---------------------------------------------------------------------------
# Costanti test
# ---------------------------------------------------------------------------

_TENANT_ID = str(uuid.uuid4())
_USER_EDITOR = str(uuid.uuid4())
_USER_VIEWER = str(uuid.uuid4())

_XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
_TEMPLATE_URL = "/api/v1/raw/excel/template"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_user(role: str, user_id: str | None = None) -> CurrentUser:
    """Crea un CurrentUser per l'override delle dipendenze.

    Args:
        role: Codice ruolo RBAC.
        user_id: UUID utente opzionale.

    Returns:
        Istanza CurrentUser.
    """
    return CurrentUser(
        sub=user_id or str(uuid.uuid4()),
        role=role,  # type: ignore[arg-type]
        tenant_id=_TENANT_ID,
        jti=str(uuid.uuid4()),
    )


def _user_override(role: str, user_id: str | None = None) -> Any:
    """Override asincrono per get_current_user.

    Args:
        role: Ruolo da iniettare.
        user_id: UUID utente opzionale.

    Returns:
        Callable asincrono che restituisce CurrentUser.
    """
    user = _make_user(role, user_id)

    async def _dep() -> CurrentUser:
        return user

    return _dep


def _noop_db() -> Any:
    """Override no-op della sessione DB.

    Returns:
        Generatore asincrono che produce un MagicMock.
    """
    async def _gen() -> Any:
        session = AsyncMock()
        session.add = MagicMock()
        session.flush = AsyncMock()
        yield session

    return _gen


def _parse_xlsx(content: bytes) -> openpyxl.Workbook:
    """Legge un workbook openpyxl dai bytes della risposta HTTP.

    Args:
        content: Bytes del file .xlsx.

    Returns:
        Workbook openpyxl aperto in sola lettura.
    """
    return openpyxl.load_workbook(io.BytesIO(content), read_only=False)


# ---------------------------------------------------------------------------
# Test 1: Editor ottiene 200 + content-type xlsx corretto
# ---------------------------------------------------------------------------


def test_template_200_editor() -> None:
    """Editor (ruolo con permesso raw_ingestions.import) deve ottenere 200."""
    app.dependency_overrides[get_current_user] = _user_override("editor", _USER_EDITOR)
    app.dependency_overrides[get_db] = _noop_db()

    with TestClient(app, raise_server_exceptions=False) as client:
        resp = client.get(_TEMPLATE_URL)

    app.dependency_overrides.clear()

    assert resp.status_code == 200, f"Atteso 200, ottenuto {resp.status_code}: {resp.text}"
    assert _XLSX_CONTENT_TYPE in resp.headers.get("content-type", ""), (
        f"Content-Type atteso xlsx, ottenuto: {resp.headers.get('content-type')}"
    )
    # Verifica che i bytes siano un XLSX valido (magic bytes PK\x03\x04)
    assert resp.content[:4] == b"PK\x03\x04", "Risposta non inizia con magic bytes XLSX"
    # Verifica Content-Disposition
    content_disp = resp.headers.get("content-disposition", "")
    assert "attachment" in content_disp, (
        f"Content-Disposition non contiene 'attachment': {content_disp}"
    )
    assert "carbontrace_template.xlsx" in content_disp, (
        f"Content-Disposition non contiene 'carbontrace_template.xlsx': {content_disp}"
    )


# ---------------------------------------------------------------------------
# Test 2: Admin ottiene 200
# ---------------------------------------------------------------------------


def test_template_200_admin() -> None:
    """Admin deve poter scaricare il template (permesso raw_ingestions.import)."""
    app.dependency_overrides[get_current_user] = _user_override("admin")
    app.dependency_overrides[get_db] = _noop_db()

    with TestClient(app, raise_server_exceptions=False) as client:
        resp = client.get(_TEMPLATE_URL)

    app.dependency_overrides.clear()

    assert resp.status_code == 200, f"Atteso 200 per admin, ottenuto {resp.status_code}"


# ---------------------------------------------------------------------------
# Test 3: Viewer ottiene 403
# ---------------------------------------------------------------------------


def test_template_403_viewer() -> None:
    """Viewer non ha il permesso raw_ingestions.import e deve ottenere 403."""
    app.dependency_overrides[get_current_user] = _user_override("viewer", _USER_VIEWER)
    app.dependency_overrides[get_db] = _noop_db()

    with TestClient(app, raise_server_exceptions=False) as client:
        resp = client.get(_TEMPLATE_URL)

    app.dependency_overrides.clear()

    assert resp.status_code == 403, (
        f"Atteso 403 per viewer, ottenuto {resp.status_code}: {resp.text}"
    )
    body = resp.json()
    assert body.get("status") == 403 or "Forbidden" in str(body), (
        f"Risposta 403 non ha il formato atteso: {body}"
    )


# ---------------------------------------------------------------------------
# Test 4: Il workbook ha esattamente 3 fogli scope + 1 _README
# ---------------------------------------------------------------------------


def test_template_sheet_count() -> None:
    """Il workbook deve avere esattamente 4 fogli: scope1, scope2, scope3, _README."""
    app.dependency_overrides[get_current_user] = _user_override("editor", _USER_EDITOR)
    app.dependency_overrides[get_db] = _noop_db()

    with TestClient(app, raise_server_exceptions=False) as client:
        resp = client.get(_TEMPLATE_URL)

    app.dependency_overrides.clear()

    assert resp.status_code == 200
    wb = _parse_xlsx(resp.content)
    sheet_names = wb.sheetnames

    assert len(sheet_names) == 4, (
        f"Attesi 4 fogli, trovati {len(sheet_names)}: {sheet_names}"
    )
    assert "scope1" in sheet_names, f"Foglio 'scope1' mancante. Fogli presenti: {sheet_names}"
    assert "scope2" in sheet_names, f"Foglio 'scope2' mancante. Fogli presenti: {sheet_names}"
    assert "scope3" in sheet_names, f"Foglio 'scope3' mancante. Fogli presenti: {sheet_names}"
    assert "_README" in sheet_names, f"Foglio '_README' mancante. Fogli presenti: {sheet_names}"


# ---------------------------------------------------------------------------
# Test 5: Ogni foglio scope ha gli header attesi in riga 1
# ---------------------------------------------------------------------------


def test_template_scope_headers() -> None:
    """Riga 1 di ogni foglio scope deve corrispondere alle colonne canoniche."""
    app.dependency_overrides[get_current_user] = _user_override("editor", _USER_EDITOR)
    app.dependency_overrides[get_db] = _noop_db()

    with TestClient(app, raise_server_exceptions=False) as client:
        resp = client.get(_TEMPLATE_URL)

    app.dependency_overrides.clear()

    assert resp.status_code == 200
    wb = _parse_xlsx(resp.content)

    expected_headers: dict[str, list[str]] = {
        "scope1": _SCOPE1_HEADERS,
        "scope2": _SCOPE2_HEADERS,
        "scope3": _SCOPE3_HEADERS,
    }

    for sheet_name, headers in expected_headers.items():
        ws = wb[sheet_name]
        row1_values = [ws.cell(row=1, column=i).value for i in range(1, len(headers) + 1)]
        assert row1_values == headers, (
            f"Foglio '{sheet_name}': header attesi {headers}, "
            f"header trovati {row1_values}"
        )


# ---------------------------------------------------------------------------
# Test 6: Gli header sono in bold (verifica stile openpyxl)
# ---------------------------------------------------------------------------


def test_template_headers_bold() -> None:
    """Le celle di intestazione (riga 1) di ogni foglio scope devono essere bold."""
    app.dependency_overrides[get_current_user] = _user_override("editor", _USER_EDITOR)
    app.dependency_overrides[get_db] = _noop_db()

    with TestClient(app, raise_server_exceptions=False) as client:
        resp = client.get(_TEMPLATE_URL)

    app.dependency_overrides.clear()

    assert resp.status_code == 200
    wb = _parse_xlsx(resp.content)

    for sheet_name in ("scope1", "scope2", "scope3"):
        ws = wb[sheet_name]
        # Verifica che almeno la prima cella header sia bold
        first_cell = ws.cell(row=1, column=1)
        assert first_cell.font is not None and first_cell.font.bold, (
            f"Foglio '{sheet_name}': la cella A1 non è in bold. "
            f"Font: {first_cell.font}"
        )
        # Verifica che tutte le celle header siano bold
        col_count = ws.max_column or 1
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value is not None:
                assert cell.font is not None and cell.font.bold, (
                    f"Foglio '{sheet_name}', colonna {col_idx} ({cell.value!r}): "
                    f"non è in bold."
                )


# ---------------------------------------------------------------------------
# Test unit: build_excel_template() restituisce bytes validi (senza HTTP)
# ---------------------------------------------------------------------------


def test_build_excel_template_returns_valid_xlsx() -> None:
    """build_excel_template() deve restituire bytes XLSX validi con 4 fogli."""
    result = build_excel_template()

    assert isinstance(result, bytes), "build_excel_template deve restituire bytes"
    assert len(result) > 0, "Il template non deve essere vuoto"
    # Magic bytes XLSX
    assert result[:4] == b"PK\x03\x04", "I bytes non iniziano con il magic XLSX"

    wb = openpyxl.load_workbook(io.BytesIO(result))
    assert set(wb.sheetnames) == {"scope1", "scope2", "scope3", "_README"}, (
        f"Fogli attesi {{scope1, scope2, scope3, _README}}, trovati: {wb.sheetnames}"
    )


def test_build_excel_template_with_known_sites() -> None:
    """build_excel_template(known_sites=...) deve completarsi senza errori."""
    extra_sites = ["SITE_A", "SITE_B", "SITE_C"]
    result = build_excel_template(known_sites=extra_sites)
    assert isinstance(result, bytes)
    assert len(result) > 0


def test_build_excel_template_example_row_count() -> None:
    """Ogni foglio scope deve avere esattamente 1 riga di esempio (riga 2)."""
    result = build_excel_template()
    wb = openpyxl.load_workbook(io.BytesIO(result))

    for sheet_name in ("scope1", "scope2", "scope3"):
        ws = wb[sheet_name]
        # La riga 1 è l'header, la riga 2 è l'esempio
        # Non ci devono essere dati dalla riga 3 in poi
        max_data_row = ws.max_row
        assert max_data_row == 2, (  # noqa: PLR2004
            f"Foglio '{sheet_name}': attese 2 righe (header + esempio), "
            f"trovate {max_data_row}"
        )
        # La riga 2 deve avere almeno un valore
        row2_values = [ws.cell(row=2, column=i).value for i in range(1, ws.max_column + 1)]
        assert any(v is not None for v in row2_values), (
            f"Foglio '{sheet_name}': la riga di esempio (riga 2) è vuota"
        )
