"""Tests for the Excel builder — verifies XLSX magic bytes and all 11 sheets (FR-27)."""

from __future__ import annotations

import io

import pytest

from ghg_tool.ui.excel.builder import REQUIRED_SHEET_NAMES, XlsxBuilder


@pytest.fixture()
def sample_report_data() -> dict:  # type: ignore[type-arg]
    """Minimal report_data fixture for Excel generation tests."""
    return {
        "anno": 2025,
        "gwp_set": "AR6",
        "emissions": [
            {
                "scope": 1, "sub_scope": "combustion", "codice_sito": "IANO",
                "anno": 2025, "tco2e": 1234.5678,
                "factor_source": "DEFRA", "factor_version": "2024",
                "gwp_set": "AR6", "methodology": "activity-based",
                "co2_tonne": 1200.0, "ch4_tco2e": 20.0, "n2o_tco2e": 14.5,
                "disclosure_notes": None,
            },
            {
                "scope": 1, "sub_scope": "process", "codice_sito": "IANO",
                "anno": 2025, "tco2e": 4115.2,
                "factor_source": "IPCC", "factor_version": "AR6",
                "gwp_set": "AR6", "methodology": "stoichiometric",
                "co2_tonne": 4115.2, "ch4_tco2e": None, "n2o_tco2e": None,
                "disclosure_notes": "Stoichiometric 0.4397 tCO2/t CaCO3",
            },
            {
                "scope": 2, "sub_scope": "LB", "codice_sito": "IANO",
                "anno": 2025, "tco2e": 500.0,
                "factor_source": "ISPRA", "factor_version": "2024",
                "gwp_set": "AR6", "methodology": "location-based",
                "disclosure_notes": None,
            },
            {
                "scope": 2, "sub_scope": "MB", "codice_sito": "IANO",
                "anno": 2025, "tco2e": 0.0,
                "factor_source": "ISPRA", "factor_version": "2024",
                "gwp_set": "AR6", "methodology": "market-based",
                "disclosure_notes": None,
            },
            {
                "scope": 3, "sub_scope": "Cat1", "codice_sito": None,
                "anno": 2025, "tco2e": 3200.0,
                "factor_source": "ecoinvent", "factor_version": "v3.10",
                "gwp_set": "AR6", "methodology": "mass-based",
                "disclosure_notes": None,
            },
        ],
        "biogenic": [
            {
                "scope": 1, "sub_scope": "biogenic", "codice_sito": "IANO",
                "anno": 2025, "co2_biogenic_tonne": 100.0,
                "co2_fossil_tonne": 1100.0,
                "factor_source": "DEFRA", "factor_version": "2024",
                "gwp_set": "AR6", "disclosure_notes": "ADR-007 memo",
            }
        ],
        "factors": [
            {
                "factor_id": "GAS_NAT_DEFRA_2024", "version": "2024",
                "substance": "Gas naturale", "scope": 1, "category": "combustion",
                "source": "DEFRA", "value": 2.042, "unit": "kgCO2e/m3",
                "gwp_set": "AR6", "valid_from": "2024-01-01",
                "applicability_note": "UK DEFRA 2024",
            }
        ],
        "dq_findings": [
            {
                "id": "aaaaaaaa-0000-0000-0000-000000000001",
                "rule_id": "DQ-CRIT-01", "severity": "CRIT",
                "resolution_status": "OPEN", "scope": 1,
                "codice_sito": "IANO", "anno": 2025,
                "metric": "facility_coverage",
                "value_observed": 6.0, "value_reference": 7.0,
                "trigger_desc": "Coverage < 7/7", "recommended_action": "Check ETL",
            }
        ],
        "audit_trail": [
            {
                "emission_id": "bbbbbbbb-0000-0000-0000-000000000001",
                "superseded_by": None, "correlation_id": "cccccccc-0000-0000-0000-000000000001",
                "scope": 1, "sub_scope": "combustion", "codice_sito": "IANO",
                "anno": 2025, "tco2e": 1234.5678,
                "factor_source": "DEFRA", "factor_version": "2024",
                "gwp_set": "AR6", "methodology": "activity-based",
                "calc_timestamp": "2026-01-01T00:00:00Z",
                "valid_from": "2026-01-01T00:00:00Z", "valid_to": None,
                "created_by": "data_steward", "reason_code": None,
            }
        ],
    }


class TestXlsxBuilder:
    def test_build_returns_bytes(self, sample_report_data: dict) -> None:  # type: ignore[type-arg]
        builder = XlsxBuilder()
        result = builder.build(sample_report_data)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_xlsx_magic_bytes(self, sample_report_data: dict) -> None:  # type: ignore[type-arg]
        """XLSX is a ZIP file: must start with PK\\x03\\x04."""
        builder = XlsxBuilder()
        result = builder.build(sample_report_data)
        assert result[:4] == b"PK\x03\x04", (
            f"Expected ZIP magic bytes PK\\x03\\x04, got {result[:4]!r}"
        )

    def test_all_11_sheets_present(self, sample_report_data: dict) -> None:  # type: ignore[type-arg]
        """All 11 required sheets must be present in the workbook."""
        import openpyxl

        builder = XlsxBuilder()
        result = builder.build(sample_report_data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        actual_sheets = wb.sheetnames
        for required in REQUIRED_SHEET_NAMES:
            assert required in actual_sheets, (
                f"Sheet '{required}' missing. Present: {actual_sheets}"
            )

    def test_summary_sheet_has_data(self, sample_report_data: dict) -> None:  # type: ignore[type-arg]
        import openpyxl

        builder = XlsxBuilder()
        result = builder.build(sample_report_data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Summary"]
        # Should have header row + at least one data row
        assert ws.max_row >= 2

    def test_biogenic_sheet_has_adr007_note(self, sample_report_data: dict) -> None:  # type: ignore[type-arg]
        import openpyxl

        builder = XlsxBuilder()
        result = builder.build(sample_report_data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Biogenic Memo"]
        # ADR-007 disclaimer must appear in the sheet
        sheet_text = " ".join(
            str(cell.value or "") for row in ws.iter_rows() for cell in row
        )
        assert "ADR-007" in sheet_text

    def test_methodology_sheet_has_content(self, sample_report_data: dict) -> None:  # type: ignore[type-arg]
        import openpyxl

        builder = XlsxBuilder()
        result = builder.build(sample_report_data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Methodology"]
        assert ws.max_row >= 3

    def test_empty_emissions_produces_valid_workbook(self) -> None:
        builder = XlsxBuilder()
        empty_data: dict = {  # type: ignore[type-arg]
            "anno": 2024, "gwp_set": "AR6",
            "emissions": [], "biogenic": [], "factors": [],
            "dq_findings": [], "audit_trail": [],
        }
        result = builder.build(empty_data)
        assert result[:4] == b"PK\x03\x04"

    def test_factor_catalog_sheet_has_data(self, sample_report_data: dict) -> None:  # type: ignore[type-arg]
        import openpyxl

        builder = XlsxBuilder()
        result = builder.build(sample_report_data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Factor Catalog"]
        assert ws.max_row >= 2  # header + at least 1 factor

    def test_dq_findings_sheet_has_data(self, sample_report_data: dict) -> None:  # type: ignore[type-arg]
        import openpyxl

        builder = XlsxBuilder()
        result = builder.build(sample_report_data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["DQ Findings"]
        assert ws.max_row >= 2

    def test_audit_trail_sheet_has_data(self, sample_report_data: dict) -> None:  # type: ignore[type-arg]
        import openpyxl

        builder = XlsxBuilder()
        result = builder.build(sample_report_data)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb["Audit Trail"]
        assert ws.max_row >= 2
