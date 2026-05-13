"""XlsxBuilder — produces the 11-sheet Excel workbook (FR-27).

Usage:
    builder = XlsxBuilder()
    xlsx_bytes = builder.build(report_data)
    assert xlsx_bytes[:4] == b"PK\\x03\\x04"  # ZIP magic

All numeric cells have factor_source, factor_version, gwp_set in adjacent
metadata columns (FR-23 transparency requirement).
Header row: bold + Okabe-Ito Blue fill (#0072B2).
Biogenic sheet explicitly flags that values are NOT in Scope totals (ADR-007).

11 sheets (FR-27):
  1. Summary
  2. Scope 1 Combustion
  3. Scope 1 Process
  4. Scope 2 LB
  5. Scope 2 MB
  6. Scope 3 Cat 1-12
  7. Biogenic Memo
  8. Factor Catalog
  9. DQ Findings
  10. Audit Trail
  11. Methodology
"""

from __future__ import annotations

import io
from typing import Any

import structlog
from openpyxl import Workbook

from ghg_tool.ui.excel.sheets import (
    write_audit_trail_sheet,
    write_biogenic_memo_sheet,
    write_dq_findings_sheet,
    write_factor_catalog_sheet,
    write_methodology_sheet,
    write_scope1_combustion_sheet,
    write_scope1_process_sheet,
    write_scope2_lb_sheet,
    write_scope2_mb_sheet,
    write_scope3_sheet,
    write_summary_sheet,
)

logger = structlog.get_logger(__name__)

# Required sheet names (used in test assertions)
REQUIRED_SHEET_NAMES = [
    "Summary",
    "Scope 1 Combustion",
    "Scope 1 Process",
    "Scope 2 LB",
    "Scope 2 MB",
    "Scope 3 Cat 1-12",
    "Biogenic Memo",
    "Factor Catalog",
    "DQ Findings",
    "Audit Trail",
    "Methodology",
]


class XlsxBuilder:
    """Builds the 11-sheet Excel workbook from pre-calculated report_data.

    No emission calculations are performed here — all values come from the
    ``report_data`` dict assembled by the export_service or data_loader.
    """

    def build(self, report_data: dict[str, Any]) -> bytes:
        """Generate the 11-sheet Excel workbook as bytes.

        Args:
            report_data: Dict with keys:
                - anno: int
                - gwp_set: str
                - emissions: list of emission row dicts
                - biogenic: list of biogenic memo dicts (ADR-007)
                - factors: list of factor catalog dicts
                - dq_findings: list of DQ finding dicts
                - audit_trail: list of audit trail dicts

        Returns:
            XLSX bytes starting with ``b'PK\\x03\\x04'`` (ZIP magic).
        """
        anno = report_data.get("anno", "?")
        gwp_set = report_data.get("gwp_set", "AR6")

        log = logger.bind(anno=anno, gwp_set=gwp_set)
        log.info("Building Excel workbook")

        wb = Workbook()
        # Remove the default blank sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        # Write all 11 sheets in order
        write_summary_sheet(wb, report_data)
        write_scope1_combustion_sheet(wb, report_data)
        write_scope1_process_sheet(wb, report_data)
        write_scope2_lb_sheet(wb, report_data)
        write_scope2_mb_sheet(wb, report_data)
        write_scope3_sheet(wb, report_data)
        write_biogenic_memo_sheet(wb, report_data)
        write_factor_catalog_sheet(wb, report_data)
        write_dq_findings_sheet(wb, report_data)
        write_audit_trail_sheet(wb, report_data)
        write_methodology_sheet(wb, report_data)

        # Verify all 11 sheets were created
        missing = [s for s in REQUIRED_SHEET_NAMES if s not in wb.sheetnames]
        if missing:
            log.warning("Missing expected sheets", missing=missing)

        # Serialise to bytes in memory
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        xlsx_bytes = buf.read()

        log.info("Excel workbook generated", size_bytes=len(xlsx_bytes),
                 sheets=len(wb.sheetnames))
        return xlsx_bytes
