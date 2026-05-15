"""Excel sheet helpers — one function per sheet (FR-27).

Each function takes an openpyxl ``Workbook``, a sheet name, data, and writes
the sheet with:
  - Bold header row with Okabe-Ito Blue fill (EXCEL_HEADER_FILL)
  - Adjacent metadata columns (factor_source, factor_version, gwp_set)
    for every numeric value (FR-23 transparency requirement)
  - Auto-width columns (heuristic)

Biogenic sheet explicitly excludes biogenic values from Scope totals (ADR-007).

Task REV-WAVE3-010: factor source strings in the Summary sheet are now built
dynamically from ``report_data["factors"]`` (the factor catalog snapshot for
the run) instead of being hardcoded as "DEFRA 2024" etc.
"""

from __future__ import annotations

from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ghg_tool.ui.streamlit_app.lib.palette import EXCEL_HEADER_FILL

# Openpyxl fill for header row (Okabe-Ito Blue: #0072B2)
_HEADER_FILL = PatternFill(fill_type="solid", fgColor=EXCEL_HEADER_FILL)
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

# CSV / Excel formula-injection prefixes (OWASP — "CSV Injection").  When a
# free-text cell value starts with one of these characters, Excel will
# interpret the cell as a formula on open. We defang by prefixing with a
# single apostrophe so the cell renders as a literal string.
_FORMULA_INJECTION_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell_value(value: Any) -> Any:
    """Defang potential Excel formula injection in string cell values.

    Numbers, dates, booleans and ``None`` pass through unchanged.  Strings
    starting with a formula-trigger character are prefixed with ``'`` so
    Excel treats them as literal text instead of evaluating them.

    Args:
        value: Raw cell value to write.

    Returns:
        The same value, or a defanged string when injection is detected.
    """
    if isinstance(value, str) and value.startswith(_FORMULA_INJECTION_PREFIXES):
        return "'" + value
    return value


def _write_header(ws: Any, columns: list[str]) -> None:
    """Write a bold header row with Okabe-Ito Blue fill.

    Args:
        ws: openpyxl Worksheet.
        columns: List of column header strings.
    """
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN
    ws.row_dimensions[1].height = 20


def _auto_width(ws: Any) -> None:
    """Set heuristic column widths based on header + first data row.

    Args:
        ws: openpyxl Worksheet.
    """
    for col in ws.iter_cols():
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 50)


# ---------------------------------------------------------------------------
# Internal helper — REV-WAVE3-010: build dynamic factor-source label
# ---------------------------------------------------------------------------

def _build_factor_source_label(factors: list[dict[str, Any]]) -> str:
    """Build a human-readable factor-source label from the catalog snapshot.

    For each published factor in ``factors`` builds a string of the form::

        "DEFRA 2024.1 (pubblicato 2024-01-15)"

    If ``published_at`` is absent (draft or legacy row), falls back to::

        "DEFRA 2024.1"

    All distinct strings are joined with a semicolon separator so multiple
    sources appear on one line.  When ``factors`` is empty the fallback
    "Vedi catalogo fattori" is returned.

    Args:
        factors: List of factor catalog dicts from ``report_data["factors"]``.

    Returns:
        A non-empty string label describing the factor sources used.
    """
    if not factors:
        return "Vedi catalogo fattori"

    seen: set[str] = set()
    parts: list[str] = []
    for f in factors:
        source = f.get("source") or ""
        version = f.get("version") or ""
        published_at = f.get("published_at") or f.get("valid_from") or ""
        # Format published_at: keep only the date part (first 10 chars)
        if published_at and len(str(published_at)) >= 10:
            pub_str = str(published_at)[:10]
            label = f"{source} {version} (pubblicato {pub_str})"
        else:
            label = f"{source} {version}".strip()
        if label and label not in seen:
            seen.add(label)
            parts.append(label)

    return "; ".join(parts) if parts else "Vedi catalogo fattori"


# ---------------------------------------------------------------------------
# Sheet 1 — Summary
# ---------------------------------------------------------------------------
def write_summary_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write the Summary sheet with scope totals and biogenic memo.

    REV-WAVE3-010: the ``factor_source`` column now contains a dynamic string
    built from ``report_data["factors"]`` (the factor catalog snapshot for
    the run) instead of a hardcoded placeholder.  The caption row below the
    data table points auditors to ``docs/methodology/factor_sources.md``.

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("Summary")
    anno = report_data.get("anno", "?")
    gwp_set = report_data.get("gwp_set", "AR6")
    emissions = report_data.get("emissions", [])
    factors_used: list[dict[str, Any]] = report_data.get("factors", [])

    # REV-WAVE3-010: build dynamic label from factor catalog snapshot
    factor_sources_label = _build_factor_source_label(factors_used)

    cols = ["Scope", "Sub-scope", "Totale tCO2e", "Anno",
            "factor_source", "factor_version", "gwp_set", "Note"]
    _write_header(ws, cols)

    scope_totals: dict[int, float] = {1: 0.0, 2: 0.0, 3: 0.0}
    scope2_lb = 0.0
    scope2_mb = 0.0
    biogenic_total = 0.0
    for r in emissions:
        scope = r.get("scope")
        tco2e = float(r.get("tco2e", 0))
        sub = r.get("sub_scope", "")
        if scope in (1, 2, 3):
            if scope == 2:
                if sub == "LB":
                    scope2_lb += tco2e
                elif sub == "MB":
                    scope2_mb += tco2e
            else:
                scope_totals[scope] += tco2e
        if sub == "biogenic":
            biogenic_total += tco2e

    # Write scope summary rows
    for scope_num, total in scope_totals.items():
        ws.append([
            f"Scope {scope_num}", "Totale",
            total, anno,
            factor_sources_label, "Vedi catalogo fattori", gwp_set,
            "Non include biogenico (ADR-007)" if scope_num == 1 else "",
        ])

    ws.append(["Scope 2", "Location-Based", scope2_lb, anno,
        factor_sources_label, "Vedi catalogo fattori", gwp_set, "",
    ])
    ws.append([
        "Scope 2", "Market-Based", scope2_mb, anno,
        factor_sources_label, "Vedi catalogo fattori", gwp_set, "",
    ])

    # Biogenic memo (NOT summed into totals per ADR-007)
    ws.append([
        "BIOGENICO (MEMO E1-7)", "ADR-007 — Non incluso in totali",
        biogenic_total, anno,
        "N/A", "N/A", gwp_set,
        "ADR-007: CO2 biogenica esclusa da Scope 1/2/3 per GHG Protocol §4.5",
    ])

    # REV-WAVE3-010: caption row pointing auditors to methodology docs
    _last_row = ws.max_row + 2
    caption_cell = ws.cell(row=_last_row, column=1)
    caption_cell.value = (
        "Fattori conformi ai dati DB. "
        "Vedi docs/methodology/factor_sources.md per i SHA-256 dei PDF sorgente."
    )
    caption_cell.font = Font(italic=True, color="FF777777")
    ws.merge_cells(
        start_row=_last_row, start_column=1,
        end_row=_last_row, end_column=len(cols),
    )

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 2 — Scope 1 Combustion
# ---------------------------------------------------------------------------
def write_scope1_combustion_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write Scope 1 combustion breakdown.

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("Scope 1 Combustion")
    emissions = [r for r in report_data.get("emissions", [])
                 if r.get("scope") == 1 and r.get("sub_scope") == "combustion"]

    cols = ["Sito", "Anno", "Sub-scope", "Combustibile",
            "tCO2e", "CO2 (t)", "CH4 (tCO2e)", "N2O (tCO2e)",
            "factor_source", "factor_version", "gwp_set", "methodology"]
    _write_header(ws, cols)

    for r in emissions:
        ws.append([
            r.get("codice_sito"), r.get("anno"), r.get("sub_scope"), "",
            r.get("tco2e", 0), r.get("co2_tonne"), r.get("ch4_tco2e"),
            r.get("n2o_tco2e"), r.get("factor_source"), r.get("factor_version"),
            r.get("gwp_set"), r.get("methodology"),
        ])
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 3 — Scope 1 Process
# ---------------------------------------------------------------------------
def write_scope1_process_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write Scope 1 process emissions (decarbonation, IANO only).

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("Scope 1 Process")
    emissions = [r for r in report_data.get("emissions", [])
                 if r.get("scope") == 1 and r.get("sub_scope") == "process"]

    cols = ["Sito", "Anno", "Sub-scope", "tCO2e", "CO2 (t)",
            "factor_source", "factor_version", "gwp_set", "methodology", "Note"]
    _write_header(ws, cols)

    for r in emissions:
        ws.append([
            r.get("codice_sito"), r.get("anno"), r.get("sub_scope"),
            r.get("tco2e", 0), r.get("co2_tonne"),
            r.get("factor_source"), r.get("factor_version"),
            r.get("gwp_set"), r.get("methodology"),
            r.get("disclosure_notes", ""),
        ])
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 4 — Scope 2 LB
# ---------------------------------------------------------------------------
def write_scope2_lb_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write Scope 2 Location-Based per facility per year.

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("Scope 2 LB")
    emissions = [r for r in report_data.get("emissions", [])
                 if r.get("scope") == 2 and r.get("sub_scope") == "LB"]

    cols = ["Sito", "Anno", "tCO2e (Location-Based)",
            "factor_source", "factor_version", "gwp_set", "methodology"]
    _write_header(ws, cols)

    for r in emissions:
        ws.append([
            r.get("codice_sito"), r.get("anno"), r.get("tco2e", 0),
            r.get("factor_source"), r.get("factor_version"),
            r.get("gwp_set"), r.get("methodology"),
        ])
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 5 — Scope 2 MB
# ---------------------------------------------------------------------------
def write_scope2_mb_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write Scope 2 Market-Based per facility per year.

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("Scope 2 MB")
    emissions = [r for r in report_data.get("emissions", [])
                 if r.get("scope") == 2 and r.get("sub_scope") == "MB"]

    cols = ["Sito", "Anno", "tCO2e (Market-Based)", "Strumento MB",
            "factor_source", "factor_version", "gwp_set", "methodology"]
    _write_header(ws, cols)

    for r in emissions:
        ws.append([
            r.get("codice_sito"), r.get("anno"), r.get("tco2e", 0),
            r.get("regulatory_stream", ""),
            r.get("factor_source"), r.get("factor_version"),
            r.get("gwp_set"), r.get("methodology"),
        ])
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 6 — Scope 3 Cat 1–12
# ---------------------------------------------------------------------------
def write_scope3_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write Scope 3 categories 1–12 per facility per year.

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("Scope 3 Cat 1-12")
    emissions = [r for r in report_data.get("emissions", [])
                 if r.get("scope") == 3]

    cols = ["Categoria", "Sub-scope", "Sito", "Anno", "tCO2e",
            "factor_source", "factor_version", "gwp_set",
            "methodology", "Note"]
    _write_header(ws, cols)

    for r in emissions:
        ws.append([
            r.get("sub_scope", "")[:4],  # e.g. "Cat1"
            r.get("sub_scope"), r.get("codice_sito"), r.get("anno"),
            r.get("tco2e", 0),
            r.get("factor_source"), r.get("factor_version"),
            r.get("gwp_set"), r.get("methodology"),
            r.get("disclosure_notes", ""),
        ])
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 7 — Biogenic Memo
# ---------------------------------------------------------------------------
def write_biogenic_memo_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write biogenic memo sheet (ADR-007 — NOT summed in totals).

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("Biogenic Memo")
    biogenic = report_data.get("biogenic", [])

    # Mandatory ADR-007 disclaimer in row 1
    ws.cell(row=1, column=1).value = (
        "ADR-007: Le emissioni biogeniche di CO2 dalla combustione di biomasse e biogas "
        "sono disclosed separatamente. NON incluse in Scope 1/2/3."
    )
    ws.cell(row=1, column=1).font = Font(bold=True, color="FFCC0000")
    ws.row_dimensions[1].height = 30

    cols = ["Sito", "Anno", "Sub-scope", "CO2 Biogenico (t)", "CO2 Fossile (t)",
            "factor_source", "factor_version", "gwp_set", "Note"]
    _write_header(ws, cols)  # writes to row 1 again — shift down

    # Rewrite header at row 2
    ws.delete_rows(1)  # remove the header we just wrote
    ws.insert_rows(1)  # make room for disclaimer
    ws.cell(row=1, column=1).value = (
        "ADR-007: CO2 Biogenica — NON inclusa in totali Scope 1/2/3 (GHG Protocol §4.5)"
    )
    ws.cell(row=1, column=1).font = Font(bold=True, color="FFCC0000")

    # Write actual header at row 2
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    row_num = 3
    for r in biogenic:
        ws.cell(row=row_num, column=1).value = r.get("codice_sito")
        ws.cell(row=row_num, column=2).value = r.get("anno")
        ws.cell(row=row_num, column=3).value = r.get("sub_scope")
        ws.cell(row=row_num, column=4).value = r.get("co2_biogenic_tonne", 0)
        ws.cell(row=row_num, column=5).value = r.get("co2_fossil_tonne", 0)
        ws.cell(row=row_num, column=6).value = r.get("factor_source")
        ws.cell(row=row_num, column=7).value = r.get("factor_version")
        ws.cell(row=row_num, column=8).value = r.get("gwp_set")
        ws.cell(row=row_num, column=9).value = r.get("disclosure_notes", "")
        row_num += 1

    if row_num == 3:  # no data
        ws.cell(row=3, column=1).value = "Nessuna emissione biogenica nel periodo."

    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 8 — Factor Catalog
# ---------------------------------------------------------------------------
def write_factor_catalog_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write a snapshot of the factor catalog used in this report.

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("Factor Catalog")
    factors = report_data.get("factors", [])

    cols = ["Factor ID", "Versione", "Sostanza", "Scope", "Categoria",
            "Fonte", "Valore", "Unità", "GWP Set", "Valido dal", "Note"]
    _write_header(ws, cols)

    for f in factors:
        ws.append([
            _safe_cell_value(f.get("factor_id")),
            _safe_cell_value(f.get("version")),
            _safe_cell_value(f.get("substance")),
            f.get("scope"),
            _safe_cell_value(f.get("category")),
            _safe_cell_value(f.get("source")),
            f.get("value"),
            _safe_cell_value(f.get("unit")),
            _safe_cell_value(f.get("gwp_set")),
            str(f.get("valid_from", "")),
            _safe_cell_value(f.get("applicability_note", "")),
        ])
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 9 — DQ Findings
# ---------------------------------------------------------------------------
def write_dq_findings_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write DQ findings current state.

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("DQ Findings")
    findings = report_data.get("dq_findings", [])

    cols = ["ID", "Regola DQ", "Severità", "Stato", "Scope", "Sito",
            "Anno", "Metrica", "Valore osservato", "Valore riferimento",
            "Descrizione trigger", "Azione raccomandata"]
    _write_header(ws, cols)

    for f in findings:
        ws.append([
            str(f.get("id", "")), f.get("rule_id"), f.get("severity"),
            f.get("resolution_status"), f.get("scope"), f.get("codice_sito"),
            f.get("anno"), f.get("metric"), f.get("value_observed"),
            f.get("value_reference"), f.get("trigger_desc"),
            f.get("recommended_action"),
        ])
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 10 — Audit Trail
# ---------------------------------------------------------------------------
def write_audit_trail_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write all corrections in period.

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("Audit Trail")
    trail = report_data.get("audit_trail", [])

    cols = ["Emission ID", "Predecessor ID", "Correlation ID",
            "Scope", "Sub-scope", "Sito", "Anno", "tCO2e",
            "Fonte Fattore", "Versione", "GWP Set", "Metodologia",
            "Calc Timestamp", "Valid From", "Valid To",
            "Utente", "Motivo correzione"]
    _write_header(ws, cols)

    for r in trail:
        ws.append([
            r.get("emission_id"), r.get("superseded_by"), r.get("correlation_id"),
            r.get("scope"), r.get("sub_scope"), r.get("codice_sito"),
            r.get("anno"), r.get("tco2e"),
            r.get("factor_source"), r.get("factor_version"),
            r.get("gwp_set"), r.get("methodology"),
            r.get("calc_timestamp"), r.get("valid_from"), r.get("valid_to"),
            r.get("created_by"), r.get("reason_code"),
        ])
    _auto_width(ws)


# ---------------------------------------------------------------------------
# Sheet 11 — Methodology
# ---------------------------------------------------------------------------
def write_methodology_sheet(wb: Workbook, report_data: dict[str, Any]) -> None:
    """Write methodology narrative (single-cell text per section).

    Args:
        wb: Target workbook.
        report_data: Report data dict.
    """
    ws = wb.create_sheet("Methodology")
    anno = report_data.get("anno", "?")
    gwp_set = report_data.get("gwp_set", "AR6")

    sections = [
        ("Framework", "GHG Protocol Corporate Standard (2004) + Scope 2 Guidance (2015) + "
         "Scope 3 Standard (2011). CSRD ESRS E1. GRI 305. ISO 14064-1:2018."),
        ("Confine organizzativo", "Controllo operativo — 7 siti italiani."),
        ("Anno base", f"2024 (consolidato). Anno corrente: {anno}."),
        ("GWP set", (
            f"{gwp_set} (IPCC AR6 GWP100: CH4=27.9, N2O=273). "
            "AR5 disponibile per EU ETS IANO."
        )),
        ("Scope 1 combustione", "Fattori DEFRA. Gas Naturale, Gasolio, Benzina."),
        ("Scope 1 processo", "Fattore stechiometrico 0.4397 tCO2/t CaCO3 (IANO, IPCC AR6)."),
        ("Scope 1 fugitive HFC", "Zero dichiarativo — impianti ad anello chiuso."),
        ("Scope 2 LB", "ISPRA Italia grid factor (primario), IEA (secondario)."),
        ("Scope 2 MB", "0 tCO2e/MWh per GO; ISPRA residual per Grid."),
        ("Scope 3", "Cat 1: ecoinvent v3.10 + EXIOBASE. Cat 3: WTT DEFRA (quantità da Σ S1). "
         "Cat 4/9: tkm × DEFRA. Cat 7: DEFRA car (FTE HR 2024=506, 2025=484)."),
        ("Biogenico (ADR-007)", "CO2 biogenica disclosed separatamente in E1-7. "
         "NON inclusa in totali S1/2/3 per GHG Protocol §4.5 e ADR-007."),
        ("Assurance", "ISAE 3000 Limited (external assurance provider)."),
    ]

    ws.cell(row=1, column=1).value = "Sezione"
    ws.cell(row=1, column=2).value = "Testo metodologico"
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    ws.cell(row=1, column=1).font = _HEADER_FONT
    ws.cell(row=1, column=2).fill = _HEADER_FILL
    ws.cell(row=1, column=2).font = _HEADER_FONT

    for row_num, (section, text) in enumerate(sections, start=2):
        ws.cell(row=row_num, column=1).value = section
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        cell = ws.cell(row=row_num, column=2)
        cell.value = text
        cell.alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80
