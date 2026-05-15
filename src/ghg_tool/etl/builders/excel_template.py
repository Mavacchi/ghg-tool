"""Generatore del template Excel per l'import bulk emissioni (FR-03 template).

Produce un workbook .xlsx con:
  - Foglio ``scope1``: header canonici + 1 riga di esempio (da scope1_combustione.csv)
  - Foglio ``scope2``: header canonici + 1 riga di esempio (da scope2_elettricita.csv)
  - Foglio ``scope3``: header canonici + 1 riga di esempio (da scope3_categorie.csv)
  - Foglio ``_README``: istruzioni in IT+EN, codice catalogo siti/categorie, GWP default

Stilizzazione minima tramite openpyxl:
  - Riga header in grassetto (Bold)
  - Fill colore tenue su header (PatternFill azzurro chiaro)
  - Larghezza colonne adattata al contenuto

Il foglio ``_README`` è prefissato con underscore perché il parser
``excel_reader.parse_workbook`` lo ignora quando cerca i fogli scope.
"""

from __future__ import annotations

import io
from typing import Final

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colonne canoniche per ogni scope (rispecchiano _REQUIRED_BY_SCOPE in
# excel_reader.py e le intestazioni dei CSV in data/raw/).
# ---------------------------------------------------------------------------

_SCOPE1_HEADERS: Final[list[str]] = [
    "Scope", "Anno", "Codice_Sito", "Categoria_S1",
    "Combustibile", "Quantità", "Unità",
    "Fonte_Dato", "Qualità_Dato", "Stato_Dato", "Note",
]

_SCOPE2_HEADERS: Final[list[str]] = [
    "Scope", "Anno", "Codice_Sito", "Voce_S2",
    "Quantità", "Unità", "Strumento_MB",
    "Fonte_Dato", "Qualità_Dato", "Stato_Dato", "Note",
]

_SCOPE3_HEADERS: Final[list[str]] = [
    "Scope", "Anno", "Categoria_S3", "Sottocategoria",
    "Metodo", "Combustibile", "Quantità", "Unità",
    "Fonte_Dato", "Qualità_Dato", "Stato_Dato", "Note",
]

# ---------------------------------------------------------------------------
# Righe di esempio: prima riga reale dai file CSV in data/raw/.
# Valori hardcoded per evitare dipendenze da filesystem a runtime.
# Fonte: data/raw/scope1_combustione.csv   riga 2
#        data/raw/scope2_elettricita.csv   riga 2
#        data/raw/scope3_categorie.csv     riga 2
# ---------------------------------------------------------------------------

_SCOPE1_EXAMPLE: Final[list[str | int | float]] = [
    1, 2024, "SASSUOLO", "Benzina_Auto",
    "BENZINA", 1349, "litri",
    "SAP", "P", "Definitivo", "",
]

_SCOPE2_EXAMPLE: Final[list[str | int | float]] = [
    2, 2024, "IANO", "EE_Acquistata_GO",
    3193698, "kWh", "GO",
    "Fattura fornitore", "P", "Definitivo", "",
]

_SCOPE3_EXAMPLE: Final[list[str | int | float]] = [
    3, 2024, 4, "Feldspati_Treno",
    "Distance-based", "", 0, "tkm",
    "Dichiarazione fornitore", "S", "Definitivo", "0 km",
]

# ---------------------------------------------------------------------------
# Stili
# ---------------------------------------------------------------------------

_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="FF2E75B6",  # blu Carbontrace
    end_color="FF2E75B6",
)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

_EXAMPLE_FONT = Font(italic=True, color="FF595959")


# ---------------------------------------------------------------------------
# Istruzioni README (IT + EN)
# ---------------------------------------------------------------------------

_README_ROWS: Final[list[tuple[str, str]]] = [
    ("IT — ISTRUZIONI", "EN — INSTRUCTIONS"),
    ("", ""),
    (
        "Questo file è il template vuoto per l'import bulk emissioni in Carbontrace.",
        "This file is the empty template for bulk emission import into Carbontrace.",
    ),
    (
        "Compila i fogli scope1, scope2, scope3 con i tuoi dati.",
        "Fill in the scope1, scope2, scope3 sheets with your data.",
    ),
    (
        "La riga 1 contiene le intestazioni: NON modificare i nomi colonna.",
        "Row 1 contains the headers: do NOT change the column names.",
    ),
    (
        "La riga 2 è un esempio commentato: puoi cancellarla o sovrascriverla.",
        "Row 2 is a commented example: you can delete or overwrite it.",
    ),
    ("", ""),
    ("--- COLONNE OBBLIGATORIE ---", "--- MANDATORY COLUMNS ---"),
    (
        "Scope: 1, 2 o 3 (numero intero).",
        "Scope: 1, 2 or 3 (integer).",
    ),
    (
        "Anno: anno fiscale di competenza (es. 2024).",
        "Anno: fiscal/reporting year (e.g. 2024).",
    ),
    (
        "Codice_Sito: codice alfanumerico del sito (Scope 1 e 2).",
        "Codice_Sito: alphanumeric site code (Scope 1 and 2).",
    ),
    (
        "Quantità: valore numerico dell'attività (non negativo).",
        "Quantità: numeric activity value (non-negative).",
    ),
    (
        "Unità: unità di misura (es. litri, kWh, Sm³, t, tkm, EUR, km).",
        "Unità: unit of measure (e.g. litri, kWh, Sm³, t, tkm, EUR, km).",
    ),
    ("", ""),
    ("--- CODICI SITO (Scope 1 e 2) ---", "--- SITE CODES (Scope 1 and 2) ---"),
    (
        "SASSUOLO, FIORANO, VIANO, VIANO_GARGOLA, CASALGRANDE, IANO, FRASSINORO",
        "SASSUOLO, FIORANO, VIANO, VIANO_GARGOLA, CASALGRANDE, IANO, FRASSINORO",
    ),
    ("", ""),
    (
        "--- CATEGORIE SCOPE 3 (Categoria_S3) ---",
        "--- SCOPE 3 CATEGORIES (Categoria_S3) ---",
    ),
    (
        "1=Beni acquistati, 2=Beni capitali, 3=Fuel & energia (WTT/T&D), "
        "4=Trasporto upstream, 5=Rifiuti, 6=Viaggi di lavoro, "
        "7=Pendolarismo dipendenti, 9=Trasporto downstream, "
        "11=Uso prodotti venduti, 12=Fine vita prodotti",
        "1=Purchased goods, 2=Capital goods, 3=Fuel & energy (WTT/T&D), "
        "4=Upstream transport, 5=Waste, 6=Business travel, "
        "7=Employee commuting, 9=Downstream transport, "
        "11=Use of sold products, 12=End-of-life products",
    ),
    ("", ""),
    ("--- QUALITÀ DATO ---", "--- DATA QUALITY ---"),
    (
        "P = Primario (misurato), S = Secondario (dichiarato fornitore), "
        "E = Stimato (proxy/modello)",
        "P = Primary (measured), S = Secondary (supplier-declared), "
        "E = Estimated (proxy/model)",
    ),
    ("", ""),
    ("--- STATO DATO ---", "--- DATA STATUS ---"),
    (
        "Definitivo = confermato, Stimato = provvisorio",
        "Definitivo = confirmed, Stimato = provisional",
    ),
    ("", ""),
    ("--- GWP SET DI DEFAULT ---", "--- DEFAULT GWP SET ---"),
    (
        "Il sistema usa AR6 (IPCC Sixth Assessment Report) come default. "
        "Puoi specificare AR4 o AR5 nella richiesta API se necessario.",
        "The system uses AR6 (IPCC Sixth Assessment Report) as default. "
        "You can specify AR4 or AR5 in the API request if needed.",
    ),
    ("", ""),
    ("--- STRUMENTO_MB (solo Scope 2) ---", "--- STRUMENTO_MB (Scope 2 only) ---"),
    (
        "GO = Garanzia d'Origine, RE100 = RE100 Power Purchase Agreement, "
        "Grid_Residual = mix residuale di rete",
        "GO = Guarantee of Origin, RE100 = RE100 Power Purchase Agreement, "
        "Grid_Residual = Residual grid mix",
    ),
    ("", ""),
    (
        "Per ulteriori informazioni contatta il tuo ESG Manager.",
        "For further information contact your ESG Manager.",
    ),
]


# ---------------------------------------------------------------------------
# Funzione pubblica
# ---------------------------------------------------------------------------


def build_excel_template(
    known_sites: list[str] | None = None,
    factor_sources: list[str] | None = None,
) -> bytes:
    """Genera il template Excel vuoto per l'import bulk emissioni.

    Crea un workbook openpyxl con 4 fogli:
      - ``scope1``: header + 1 riga esempio Scope 1
      - ``scope2``: header + 1 riga esempio Scope 2
      - ``scope3``: header + 1 riga esempio Scope 3
      - ``_README``: istruzioni IT/EN, codici siti/categorie, GWP default

    I nomi dei fogli scope sono in minuscolo per corrispondere agli alias
    riconosciuti da ``excel_reader._SCOPE_SHEET_ALIASES``.

    Args:
        known_sites: Lista opzionale di codici sito da mostrare nel README.
            Se None usa i siti di default del progetto.
        factor_sources: Lista opzionale di fonti fattori da mostrare.
            Non usata nella generazione corrente; riservata per estensioni future.

    Returns:
        Bytes del file .xlsx pronto per essere inviato come risposta HTTP.
    """
    wb = openpyxl.Workbook()

    # Rimuovi il foglio di default creato da openpyxl
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)

    _add_scope_sheet(wb, "scope1", _SCOPE1_HEADERS, _SCOPE1_EXAMPLE)
    _add_scope_sheet(wb, "scope2", _SCOPE2_HEADERS, _SCOPE2_EXAMPLE)
    _add_scope_sheet(wb, "scope3", _SCOPE3_HEADERS, _SCOPE3_EXAMPLE)
    _add_readme_sheet(wb, known_sites=known_sites)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# Helpers interni
# ---------------------------------------------------------------------------


def _add_scope_sheet(
    wb: openpyxl.Workbook,
    sheet_name: str,
    headers: list[str],
    example_row: list[str | int | float],
) -> None:
    """Aggiunge un foglio scope al workbook con header stilizzati e riga esempio.

    Args:
        wb: Workbook openpyxl su cui operare.
        sheet_name: Nome del foglio (es. "scope1").
        headers: Lista nomi colonna da scrivere in riga 1.
        example_row: Valori di esempio da scrivere in riga 2 (corsivo, grigio).
    """
    ws = wb.create_sheet(title=sheet_name)

    # --- Riga 1: header ---
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT

    # --- Riga 2: esempio ---
    for col_idx, value in enumerate(example_row, start=1):
        cell = ws.cell(row=1 + 1, column=col_idx, value=value)
        cell.font = _EXAMPLE_FONT

    # --- Larghezza colonne adattata (min 12, max 40 caratteri) ---
    for col_idx, header in enumerate(headers, start=1):
        example_val = str(example_row[col_idx - 1]) if col_idx <= len(example_row) else ""
        col_width = max(len(header), len(example_val), 12)
        col_width = min(col_width + 2, 40)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # Blocca la riga header
    ws.freeze_panes = "A2"


def _add_readme_sheet(
    wb: openpyxl.Workbook,
    known_sites: list[str] | None = None,
) -> None:
    """Aggiunge il foglio ``_README`` con istruzioni bilingue.

    Il prefisso underscore garantisce che ``excel_reader.parse_workbook``
    ignori questo foglio (non corrisponde ad alcun alias scope).

    Args:
        wb: Workbook openpyxl su cui operare.
        known_sites: Lista opzionale di codici sito aggiuntivi da includere.
    """
    ws = wb.create_sheet(title="_README")

    # Intestazione del foglio
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 80

    header_cell_a = ws.cell(row=1, column=1, value="CARBONTRACE — Modello Import Excel")
    header_cell_a.font = Font(bold=True, size=14, color="FF2E75B6")
    ws.cell(row=1, column=2, value="CARBONTRACE — Excel Import Template")

    ws.cell(row=2, column=1, value="")

    row_num = 3
    for it_text, en_text in _README_ROWS:
        cell_it = ws.cell(row=row_num, column=1, value=it_text)
        cell_en = ws.cell(row=row_num, column=2, value=en_text)
        if it_text.startswith("---") and it_text.endswith("---"):
            cell_it.font = Font(bold=True)
            cell_en.font = Font(bold=True)
        row_num += 1

    # Aggiungi siti personalizzati se forniti
    if known_sites:
        ws.cell(row=row_num, column=1, value="")
        row_num += 1
        sites_cell = ws.cell(
            row=row_num, column=1,
            value=f"Siti aggiuntivi configurati: {', '.join(sorted(known_sites))}",
        )
        sites_cell.font = Font(italic=True)
        ws.cell(
            row=row_num, column=2,
            value=f"Additional configured sites: {', '.join(sorted(known_sites))}",
        ).font = Font(italic=True)
